#!/usr/bin/env python3
"""Generate blinded annotation workbooks for human double-coding.

Why blinded
-----------
The stratified sample written by ``03_classify.py`` carries the model's own
predictions in ``model_is_scientific`` / ``model_is_empirical`` /
``model_methodology``. Handing that file to a coder would destroy the thing it
exists to measure: a coder who can see the model's answer anchors on it, and
the resulting kappa measures compliance rather than agreement. This script
therefore strips every model column and writes coder-facing workbooks that
contain only what a human needs to make the judgement, plus a hidden key file
that is re-joined by ``08_score_coding.py`` at scoring time.

Design decisions
----------------
* **One workbook per coder**, so neither can see the other's codes.
* **Identical row order**, so scoring is a positional join on ``sred_id`` with
  no re-sorting.
* **Dropdown validation on every coded field**, because free-typed labels
  ("Quant", "quantitative ", "QUANT") are the single largest source of
  avoidable disagreement in hand-coding, and they inflate the apparent
  disagreement rate without any real disagreement occurring.
* **Two batches.** Batch 1 (30 records) is the calibration set: both coders do
  it first, then meet, resolve every difference, and record the resulting rules
  in the decision log. Batch 2 (270 records) is coded under the settled
  codebook. Batch 1 is re-coded afterwards so the reported kappa covers all
  300 under one set of rules.
* **A `flag` column.** A coder who is unsure should say so rather than guess;
  flagged records go to adjudication regardless of whether the two coders
  happened to agree.

Usage
-----
    python scripts/07_make_coding_workbooks.py
    python scripts/07_make_coding_workbooks.py --coders A B C --calibration 40
"""

from __future__ import annotations

import argparse
import logging
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

ROOT = Path(__file__).resolve().parents[1]
INTERIM = ROOT / "data" / "interim"
CODING = ROOT / "data" / "coding"

(ROOT / "logs").mkdir(parents=True, exist_ok=True)
logging.basicConfig(level=logging.INFO, format="%(levelname)-7s %(message)s")
log = logging.getLogger("coding")

FONT = "Arial"

# Columns that must never reach a coder.
BLINDED = ["model_is_scientific", "model_is_empirical", "model_methodology",
           "cls_backend", "cls_confidence", "prevention_level", "sdoh_focus",
           "human_is_scientific", "human_is_empirical", "human_methodology",
           "human_prevention_level", "human_sdoh_focus",
           "coder_initials", "coder_notes"]

# (column header, dropdown options, width, required?)
QUESTIONS = [
    ("Q1_communication_type",
     ["scientific", "other_scholarly"], 20, True),
    ("Q2_empirical_status",
     ["empirical", "non_empirical", "cannot_tell"], 20, True),
    ("Q3_methodology",
     ["quantitative", "qualitative", "mixed", "review", "not_applicable"], 20, True),
    ("Q4_prevention_level",
     ["universal", "selective", "indicated", "treatment", "postvention",
      "not_applicable"], 20, False),
    ("Q5_sdoh_addressed",
     ["yes", "no"], 18, False),
]

HEADER_FILL = PatternFill("solid", fgColor="8A2B2B")
INPUT_FILL = PatternFill("solid", fgColor="FFF9E0")
REF_FILL = PatternFill("solid", fgColor="F2F2F2")
CAL_FILL = PatternFill("solid", fgColor="E8F1E8")
THIN = Side(style="thin", color="D0D0D0")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def instructions_sheet(wb: Workbook, coder: str, n_cal: int, n_total: int) -> None:
    ws = wb.create_sheet("START HERE", 0)
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 104

    rows = [
        ("h1", f"SRED human coding, Coder {coder}"),
        ("p", "Read this tab before you open the Coding tab. It takes five minutes and "
              "will save you an hour."),
        ("", ""),
        ("h2", "What you are doing and why"),
        ("p", "A classifier has already labelled about 97,000 suicide research abstracts "
              "by what kind of communication they are, whether they report data, and what "
              "method they used. Your job is to code a 300-record sample by hand, without "
              "seeing what the classifier decided, so that we can measure how far the two "
              "agree. That agreement statistic goes in the paper. If you can see the "
              "machine's answer you cannot produce an independent one, which is why this "
              "workbook does not contain it."),
        ("p", "Another person is coding the identical 300 records separately. You will not "
              "see their answers either, and they will not see yours. Disagreement between "
              "you is expected and useful, not a failure. Do not try to guess what they "
              "would say."),
        ("", ""),
        ("h2", "Order of work"),
        ("p", f"BATCH 1 ({n_cal} records, shaded green) is the calibration set. Code these "
              "first, then stop. The team meets, compares, and settles any rule that turned "
              "out to be ambiguous. Do not start Batch 2 before that meeting."),
        ("p", f"BATCH 2 ({n_total - n_cal} records) is coded after the meeting, under the "
              "settled rules. You will then re-code Batch 1 so that all 300 records are "
              "coded under one consistent set of rules."),
        ("", ""),
        ("h2", "How to code one record"),
        ("p", "Read the title and abstract. Nothing else. Do not look the paper up, do not "
              "read the full text, do not search for the authors. The classifier only sees "
              "the abstract, so a fair comparison means you only see the abstract too."),
        ("p", "Answer Q1, Q2, and Q3 for every record. Q4 and Q5 are secondary; answer them "
              "when the abstract makes the answer clear and leave them blank when it does not."),
        ("p", "Use the dropdowns. Do not type values by hand: a stray capital or trailing "
              "space registers as a disagreement when there is none."),
        ("p", "Q3 only applies when Q2 is 'empirical'. If Q2 is 'non_empirical' or "
              "'cannot_tell', set Q3 to 'not_applicable'."),
        ("", ""),
        ("h2", "When you are unsure"),
        ("p", "Set FLAG to 'yes' and write what the difficulty was in NOTES. A flagged "
              "record goes to adjudication whether or not the two coders happened to land "
              "on the same answer. Flagging is not a failure and there is no quota; an "
              "honest flag is worth more to us than a confident guess."),
        ("p", "Do not leave Q1 to Q3 blank. Choose the best available answer and flag it. A "
              "blank is not the same as 'cannot_tell' and cannot be scored."),
        ("", ""),
        ("h2", "Worked example"),
        ("p", "TITLE: Safety planning after an emergency department visit for self-harm: a "
              "randomised controlled trial"),
        ("p", "ABSTRACT: Background: ... Methods: We randomly assigned 312 patients ... "
              "Results: The intervention group showed a lower rate of repeat attendance "
              "(OR 0.62, 95% CI 0.41-0.94) ..."),
        ("code", "Q1 = scientific   Q2 = empirical   Q3 = quantitative   "
                 "Q4 = indicated   Q5 = no   FLAG = no"),
        ("p", "Q3 is quantitative because the abstract reports a randomised design and a "
              "statistical estimate. Q4 is indicated because safety planning is delivered "
              "to people already identified as at risk by a self-harm presentation, rather "
              "than to a whole population. Q5 is no because no social determinant is "
              "examined as an exposure or a target."),
        ("", ""),
        ("h2", "Pace and honesty"),
        ("p", "Expect two to four minutes per record once you are warmed up, so roughly "
              "twelve to fifteen hours in total across both batches. Do not code for more "
              "than about ninety minutes at a stretch; accuracy falls off well before "
              "boredom announces itself. There is no reward for finishing early and a real "
              "cost to rushing, because your codes are the standard everything else is "
              "measured against."),
        ("p", "If you realise partway through that you have been applying a rule "
              "inconsistently, say so. We would far rather re-code a batch than publish a "
              "reliability figure we do not believe."),
        ("", ""),
        ("h2", "Full definitions"),
        ("p", "The Codebook tab in this workbook has the decision rule for every category, "
              "with the boundary cases that actually come up. Read it once before Batch 1 "
              "and keep it open while you code."),
    ]

    r = 2
    for kind, text in rows:
        c = ws.cell(row=r, column=2, value=text)
        if kind == "h1":
            c.font = Font(name=FONT, size=16, bold=True, color="8A2B2B")
        elif kind == "h2":
            c.font = Font(name=FONT, size=12, bold=True, color="8A2B2B")
        elif kind == "code":
            c.font = Font(name="Courier New", size=10, bold=True)
            c.fill = INPUT_FILL
        else:
            c.font = Font(name=FONT, size=10.5)
        c.alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[r].height = None if kind in ("h1", "h2", "") else max(
            15, 13 * (len(text) // 95 + 1))
        r += 1


def codebook_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet("Codebook")
    ws.sheet_view.showGridLines = False
    for col, w in (("A", 3), ("B", 24), ("C", 82)):
        ws.column_dimensions[col].width = w

    blocks: list[tuple[str, list[tuple[str, str]]]] = [
        ("Q1  Communication type", [
            ("scientific",
             "Presents original findings, a theoretical framework, a systematic review, or "
             "a methodological contribution. Empirical papers, conceptual papers, and "
             "formal reviews are all scientific."),
            ("other_scholarly",
             "Editorials, commentaries, letters to the editor, book reviews, news items, "
             "obituaries, corrections, conference matter. Serves a professional function "
             "but makes no original scientific contribution."),
            ("Boundary",
             "An invited commentary that develops a substantive argument is still "
             "other_scholarly if it is presented as commentary on another paper. A short "
             "report with data is scientific however brief."),
        ]),
        ("Q2  Empirical status", [
            ("empirical",
             "The abstract describes collecting or analysing data to answer a question. "
             "Original or secondary data both count. Formal evidence syntheses "
             "(systematic review, meta-analysis, scoping review) count as empirical."),
            ("non_empirical",
             "Theoretical or conceptual work, narrative discussion of a literature with no "
             "systematic method, practice commentary, or methodological argument that "
             "analyses no data."),
            ("cannot_tell",
             "The abstract genuinely does not say. Use this sparingly; it is for abstracts "
             "that describe a topic without describing what was done, not for abstracts "
             "that are merely terse."),
            ("Boundary",
             "A single case report IS empirical: one case is still data. A 'review of the "
             "literature' with no search strategy is non_empirical. A protocol paper "
             "describing a study not yet conducted is non_empirical."),
        ]),
        ("Q3  Methodology  (only when Q2 = empirical)", [
            ("quantitative",
             "Numeric data analysed statistically. Trials, cohort and case-control "
             "studies, surveys with statistics, registry analyses, psychometrics, "
             "modelling."),
            ("qualitative",
             "Textual, visual, or observational data analysed interpretively. Interviews, "
             "focus groups, thematic or content analysis, ethnography, case reports."),
            ("mixed",
             "Genuine integration of both traditions in one study. Requires that both "
             "strands be described. A survey with one open-ended question is not mixed "
             "methods."),
            ("review",
             "Systematic review, meta-analysis, or scoping review. The tell is a described "
             "search strategy or explicit inclusion criteria, not the word 'review'."),
            ("not_applicable",
             "Q2 was non_empirical or cannot_tell."),
            ("Boundary",
             "A meta-analysis is 'review', not 'quantitative', even though it reports "
             "pooled statistics. A qualitative study reporting participant counts is still "
             "qualitative. When quantitative and qualitative elements are both present but "
             "only one is analysed, code the one analysed."),
        ]),
        ("Q4  Prevention level  (secondary; blank if unclear)", [
            ("universal",
             "Delivered to a whole population regardless of risk. Media guidelines, means "
             "restriction legislation, population awareness campaigns, school-wide programmes."),
            ("selective",
             "Targeted at a group at elevated risk but not currently symptomatic. "
             "Gatekeeper training, programmes for veterans or bereaved people as a class."),
            ("indicated",
             "Targeted at individuals showing risk. Safety planning, crisis lines, "
             "follow-up contact after an attempt, screening-triggered intervention."),
            ("treatment",
             "Clinical treatment of people in care. Psychotherapy, pharmacotherapy, "
             "inpatient care, ketamine, DBT, CBT."),
            ("postvention",
             "Support after a suicide has occurred. Bereavement support, contagion response."),
            ("not_applicable",
             "Descriptive, epidemiological, aetiological, or measurement work that does not "
             "sit on the prevention continuum. This is the majority; do not force a level."),
        ]),
        ("Q5  Social determinant addressed  (secondary; blank if unclear)", [
            ("yes",
             "The study examines a social determinant as an exposure, a moderator, or an "
             "intervention target: economic conditions, education, healthcare access, "
             "neighbourhood, social connection, discrimination, housing, food insecurity, "
             "incarceration, immigration status, firearm or means access, digital and "
             "social media environment."),
            ("no",
             "No social determinant is examined. Merely adjusting for income or education "
             "as a covariate does not count; the determinant has to be part of the "
             "question, not part of the control set."),
        ]),
    ]

    r = 2
    ws.cell(row=r, column=2, value="SRED codebook").font = Font(
        name=FONT, size=16, bold=True, color="8A2B2B")
    r += 2
    for heading, items in blocks:
        c = ws.cell(row=r, column=2, value=heading)
        c.font = Font(name=FONT, size=12, bold=True, color="8A2B2B")
        r += 1
        for label, text in items:
            lc = ws.cell(row=r, column=2, value=label)
            lc.font = Font(name=FONT, size=10, bold=True,
                           color="B02020" if label == "Boundary" else "000000")
            lc.alignment = Alignment(vertical="top")
            tc = ws.cell(row=r, column=3, value=text)
            tc.font = Font(name=FONT, size=10)
            tc.alignment = Alignment(wrap_text=True, vertical="top")
            ws.row_dimensions[r].height = max(15, 13 * (len(text) // 88 + 1))
            r += 1
        r += 1


def coding_sheet(wb: Workbook, df: pd.DataFrame, n_cal: int) -> None:
    ws = wb.create_sheet("Coding")
    ws.freeze_panes = "A2"

    ref_cols = [("row", 6), ("batch", 13), ("sred_id", 26), ("year", 7),
                ("journal", 30), ("title", 52), ("abstract", 95)]
    q_cols = [(name, w) for name, _opts, w, _req in QUESTIONS]
    tail_cols = [("FLAG", 9), ("NOTES", 40)]
    headers = [h for h, _ in ref_cols] + [h for h, _ in q_cols] + [h for h, _ in tail_cols]

    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=1, column=i, value=h)
        c.font = Font(name=FONT, size=10, bold=True, color="FFFFFF")
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = BORDER
    ws.row_dimensions[1].height = 30

    widths = [w for _, w in ref_cols] + [w for _, w in q_cols] + [w for _, w in tail_cols]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    n_ref = len(ref_cols)
    for j, (_name, opts, _w, _req) in enumerate(QUESTIONS):
        col = get_column_letter(n_ref + 1 + j)
        dv = DataValidation(type="list", formula1='"' + ",".join(opts) + '"',
                            allow_blank=True, showDropDown=False)
        dv.error = "Choose a value from the dropdown."
        dv.errorTitle = "Not a valid code"
        dv.prompt = "  |  ".join(opts)
        dv.promptTitle = _name
        ws.add_data_validation(dv)
        dv.add(f"{col}2:{col}{len(df) + 1}")

    flag_col = get_column_letter(n_ref + len(QUESTIONS) + 1)
    dvf = DataValidation(type="list", formula1='"yes,no"', allow_blank=True,
                         showDropDown=False)
    ws.add_data_validation(dvf)
    dvf.add(f"{flag_col}2:{flag_col}{len(df) + 1}")

    for i, rec in enumerate(df.itertuples(index=False), start=2):
        batch = "1-calibration" if i - 1 <= n_cal else "2-main"
        vals = [i - 1, batch, rec.sred_id, rec.year, rec.journal_canonical,
                rec.title, rec.abstract]
        for j, v in enumerate(vals, start=1):
            c = ws.cell(row=i, column=j, value=v)
            c.font = Font(name=FONT, size=9.5)
            c.alignment = Alignment(wrap_text=(j >= 5), vertical="top")
            c.fill = CAL_FILL if batch.startswith("1") else REF_FILL
            c.border = BORDER
        for j in range(n_ref + 1, len(headers) + 1):
            c = ws.cell(row=i, column=j)
            c.fill = INPUT_FILL
            c.font = Font(name=FONT, size=9.5)
            c.alignment = Alignment(vertical="top", wrap_text=(j == len(headers)))
            c.border = BORDER
        ws.row_dimensions[i].height = 78

    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(df) + 1}"


def build(df: pd.DataFrame, coder: str, n_cal: int, out_dir: Path) -> Path:
    wb = Workbook()
    wb.remove(wb.active)
    instructions_sheet(wb, coder, n_cal, len(df))
    codebook_sheet(wb)
    coding_sheet(wb, df, n_cal)
    out_dir.mkdir(parents=True, exist_ok=True)
    path = out_dir / f"SRED_coding_Coder{coder}.xlsx"
    wb.save(path)
    log.info("wrote %s (%d records, %d calibration)", path.name, len(df), n_cal)
    return path


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--coders", nargs="+", default=["A", "B"])
    ap.add_argument("--calibration", type=int, default=30)
    ap.add_argument("--src", default=str(INTERIM / "human_coding_template.csv"))
    ap.add_argument("--seed", type=int, default=20260727)
    args = ap.parse_args()

    src = Path(args.src)
    if not src.exists():
        log.error("missing %s - run scripts/03_classify.py first", src)
        return 1

    df = pd.read_csv(src)
    log.info("loaded %d records, %d columns", len(df), len(df.columns))

    # The key file keeps the model predictions out of the coders' hands while
    # preserving them for scoring. Both are written from the same row order.
    key_cols = [c for c in ["sred_id", "model_is_scientific", "model_is_empirical",
                            "model_methodology", "cls_backend", "cls_confidence",
                            "prevention_level", "sdoh_focus"] if c in df.columns]
    CODING.mkdir(parents=True, exist_ok=True)
    key_path = CODING / "_scoring_key_DO_NOT_SHARE_WITH_CODERS.csv"

    # Shuffle once so neither the stratification nor the model's confidence
    # ordering leaks structure into the coding sequence.
    df = df.sample(frac=1.0, random_state=args.seed).reset_index(drop=True)
    df[key_cols].to_csv(key_path, index=False)
    log.info("wrote %s (%d columns withheld from coders)", key_path.name, len(key_cols))

    coder_df = df.drop(columns=[c for c in BLINDED if c in df.columns])
    leaked = [c for c in coder_df.columns if c.startswith(("model_", "cls_", "human_"))]
    if leaked:
        log.error("BLINDING FAILURE, these columns would reach coders: %s", leaked)
        return 1
    log.info("blinding check passed; coder columns: %s", list(coder_df.columns))

    for coder in args.coders:
        build(coder_df, coder, args.calibration, CODING)

    log.info("done. Send only the SRED_coding_Coder*.xlsx files to coders; "
             "keep %s.", key_path.name)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
